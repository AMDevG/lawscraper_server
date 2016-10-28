import sys
import os
import time

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pcos_site.settings')

from django.shortcuts import render
from django.http import HttpResponse
from django.core.mail import send_mail, EmailMessage

from datetime import date
import re
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl import Workbook

from selenium import webdriver
from selenium.webdriver.support.ui import Select
from pyvirtualdisplay import Display

def getIDs():
    start = time.time()
    search_ids = []

    day_today = date.today()
    day_today = str(day_today)
    year = str(day_today[0:4])
    mo = str(day_today[5:7])
    day = str(day_today[8:11])

    #day_today = mo+"/"+day+"/"+year
    day_today = "10/27/2016"

    display = Display(visible=0, size=(800,600))
    display.start()
    driver = webdriver.Firefox()

    driver.get('http://www.pcsoweb.com/InmateBooking/')
    date_input = driver.find_element_by_id("txtBookingDate")
    search_button = driver.find_element_by_id("btnSearch")
    page_size = Select(driver.find_element_by_id("drpPageSize"))

    date_input.send_keys(day_today)
    page_size.select_by_value('100')
    search_button.click()

    response = driver.page_source
    soup = BeautifulSoup(response)
    inmate_numbers = soup.findAll("span", {"id" : re.compile('lblJMSNumber.*')})

    for num in inmate_numbers:
    	search_ids.append(num.text)

    driver.quit()
    display.stop()

    stop = time.time()
    length = int(stop-start)

    print("Get IDs ran for seconds: ", length)

    return search_ids

def get_id_detail(test_id):

    display = Display(visible=0, size=(800,600))
    display.start()
    driver = webdriver.Firefox()

    test_id = test_id
    base_url = 'http://www.pcsoweb.com/inmatebooking/SubjectResults.aspx?id='
    target_url = base_url + test_id

    print("Calling URL and returning")
    driver.get(target_url)
    response = driver.page_source

    driver.quit()
    display.stop()

    return response

def parseTarget():
    master_data = {}

    print("Entering parse target")
    search_ids = getIDs()
    start = time.time()

    ######### TEST CODE #######
    test_ids = []
    test_ids.append("1698136")

    print("test id is ")


    print("Starting to process # ids", len(test_ids))

    for id_number in search_ids:    #use testids for test code
        detail_date = {}
    	results = []
    	charge_rows = []
    	charge_data = []


        target_html = get_id_detail(id_number)
    	soup = BeautifulSoup(target_html)

    	name = soup.find("span", {"id" : 'lblName1'}).text
    	print(name)
    	docket = soup.find("span", {"id" : 'lblDocket1'}).text
    	arrest_date = soup.find("span", {"id" : 'lblArrestDate1'}).text
    	agency = soup.find("span", {"id" : 'lblAgency'}).text
    	address = soup.find("span", {"id" : 'lblAddress'}).text
    	city = soup.find("span", {"id" : 'lblCity'}).text
    	state = soup.find("span", {"id" : 'lblState'}).text
    	zipcode = soup.find("span", {"id" : 'lblZipCode'}).text
    	race = soup.find("span", {"id" : 'lblRace1'}).text
    	sex = soup.find("span", {"id" : 'lblSex1'}).text
    	dob = soup.find("span", {"id" : 'lblDOB1'}).text
    	pob = soup.find("span", {"id" : 'lblPOB'}).text
    	arrest_age = soup.find("span", {"id" : 'lblArrestAge'}).text
    	eyes = soup.find("span", {"id" : 'lblEyes'}).text
    	hair = soup.find("span", {"id" : 'lblHair'}).text
    	complexion = soup.find("span", {"id" : 'lblComplexion'}).text
    	height = soup.find("span", {"id" : 'lblHeight'}).text
    	weight = soup.find("span", {"id" : 'lblWeight'}).text
    	markings = soup.find("span", {"id" : 'lblSMT'}).text
    	cell_location = soup.find("span", {"id" : 'CellLocation'}).text
    	account_balance = soup.find("span", {"id" : 'lblAccountBalance'}).text
    	spin = soup.find("span", {"id" : 'lblSPIN'}).text
    	booking_type = soup.find("span", {"id" : 'lblBookingType'}).text
    	alias = soup.find("span", {"id" : 'lblAKA'}).text

        detail_date['name'] = name
        detail_date['docket'] = docket
        detail_date['arrest_date'] = arrest_date
        detail_date['agency'] = agency
        detail_date['address'] = address
        detail_date['city'] = city
        detail_date['state'] = state
        detail_date['zipcode'] = zipcode
        detail_date['race'] = race
        detail_date['sex'] = sex
        detail_date['dob'] = dob
        detail_date['pob'] = pob
        detail_date['arrest_age'] = arrest_age
        detail_date['eyes'] = eyes
        detail_date['hair'] = hair
        detail_date['complexion'] = complexion
        detail_date['height'] = height
        detail_date['weight'] = weight
        detail_date['markings'] = markings
        detail_date['cell_location'] = cell_location
        detail_date['account_balance'] = account_balance
        detail_date['spin'] = spin
        detail_date['booking_type'] = booking_type
        detail_date['alias'] = alias


    	#Charge table is separate from other booking data

    	print("getting charge table")

    	charge_table = soup.find("table", id='tblCharges')
    	rows = charge_table.findAll('td')
        #### CHECK LENGTH OF CHARGE DATA IF GREATER THAN 19 MULTIPLE CHARGES
    	for row in rows:
    		charge_data.append(row.text)
    		print(row)
        print("Here are the items in charge data: ")
        # for item in charge_data:
        #     print(str(item))
    	#Iterates over every other because of headers
    	for i in range(1, len(charge_data),2):
            if i == 1:
                print("In charge_table counter is ", i)
                detail_date['charge_number'] = charge_data[i]
                print("In charge_table DATA IS ", charge_data[i])
            elif i == 3:
            	detail_date['agency_report_number'] = charge_data[i]
                print("In charge_table counter is ", i)
                print("In charge_table DATA IS ", charge_data[i])

            elif i == 5:
            	detail_date['offense'] = charge_data[i]
                print("In charge_table counter is ", i)
                print("In charge_table DATA IS ", charge_data[i])
            elif i == 7:
            	detail_date['statute'] = charge_data[i]
                print("In charge_table counter is ", i)
                print("In charge_table DATA IS ", charge_data[i])
            elif i == 9:
                detail_date['case_number'] = charge_data[i]
                print("In charge_table counter is ", i)
                print("In charge_table DATA IS ", charge_data[i])
            elif i == 11:
            	detail_date['bond_assessed'] = charge_data[i]
                print("In charge_table counter is ", i)
                print("In charge_table DATA IS ", charge_data[i])
            elif i == 13:
            	detail_date['bond_amount_due'] = charge_data[i]
                print("In charge_table counter is ", i)
                print("In charge_table DATA IS ", charge_data[i])
            elif i == 15:
            	detail_date['charge_status'] = charge_data[i]
                print("In charge_table counter is ", i)
                print("In charge_table DATA IS ", charge_data[i])
            elif i == 17:
            	detail_date['arrest_type'] = charge_data[i]
                print("In charge_table counter is ", i)
                print("In charge_table DATA IS ", charge_data[i])
            elif i == 19:
                print("In charge_table counter is ", i)
            	detail_date['obts'] = charge_data[i]
                print("OBTS is ", charge_data[i])

        master_data[docket] = detail_date

    stop = time.time()
    length = int(stop - start)
    print("It took this long to process IDs in parsetarget : ", length)
    return master_data

def write_to_excel():
    target_wb = Workbook()
    print("Initial call to write_excel")
### Data ROWS NEEDS TO BE DYNAMIC
    data_rows = ['Name', 'Docket', 'Arrest Date', 'Agency',
    			 'Address', 'City','State','Zipcode','Race',
    			 'Sex', 'Date of Birth', 'Place of Birth', 'Arrest Age',
    			 'Eye Color', 'Hair Color', 'Complexion', 'Height',
    			 'Weight', 'Markings', 'Cell Location', 'Account Balance', 'SPIN', 'Booking Type',
    			 'Alias', 'Charge Number', 'Agency Report Number', 'Offense',
    			 'Statute', 'Case Number', 'Bond Assessed', 'Bond Amount Due',
    			 'Charge Status', 'Arrest Type', 'OBTS',"RANDOM_STRING"]

    print("Length of data headers is", len(data_rows))
    print("Calling parsetarget...")

    detail_data  = parseTarget()
    start = time.time()

    print("starting to process detail data")
    print("number of keys in detail data is ", len(detail_data.keys()))

    key_count = 0
    for key in detail_data.keys():

        new_dic = detail_data[key]
        print("here are the keys in detail_data: ")
        for key in new_dic.keys():
            print(key)
            key_count += 1
        break
    print(key_count)

    for key in detail_data.keys():
        target_ws = target_wb.create_sheet()
        target_ws.title = key
    	target_ws.cell(row = 1, column = 1).value = data_rows[0]

        for i in range(1,len(data_rows)):
            target_ws.cell(row = i + 1, column = 1 ).value = data_rows[i]
            if i == 1:
                try:
            	    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['name']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 2:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['docket']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 3:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['arrest_date']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 4:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['agency']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 5:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['address']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 6:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['city']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 7:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['state']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 8:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['zipcode']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 9:
                try:
            	    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['race']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 10:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['sex']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 11:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['dob']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 12:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['pob']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 13:
                try:
                	target_ws.cell(row = i , column = 2 ).value = detail_data[key]['arrest_age']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 14:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['eyes']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 15:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['hair']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 16:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['complexion']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 17:
                try:
            	    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['height']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 18:
                try:
            	    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['weight']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 19:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['markings']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 20:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['cell_location']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 21:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['account_balance']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 22:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['spin']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 23:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['booking_type']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 24:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['alias']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 25:
                try:
                    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['charge_number']
                except KeyError:
                    target_ws.cell(row = i, column = 2 ).value = "1"
            elif i == 26:
                try:
                    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['agency_report_number']
                except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 27:
                try:
                    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['offense']
                except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 28:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['statute']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 29:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['case_number']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 30:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['bond_assessed']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 31:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['bond_amount_due']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 32:
                try:
                	target_ws.cell(row = i, column = 2 ).value = detail_data[key]['charge_status']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 33:
                try:
                    print("Counter at ", i)
                    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['arrest_type']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"
            elif i == 34:
                try:
                    print("Counter at ", i)
                    print("Writing OBTS!")
                    target_ws.cell(row = i, column = 2 ).value = detail_data[key]['obts']
            	except:
            	    target_ws.cell(row = i, column = 2 ).value = "-"

    stop = time.time()
    length = int(stop-start)
    print("Writing portion of writeXL took  ", length)

    workbook_name = "Booking Statement Report.xlsx"
    target_wb.save("/home/lawscraper/reports/"+workbook_name)
    email_attachment()

def email_attachment():
    mail = EmailMessage("New Booking Report Statement", "testemail", 'bprecosheet@gmail.com', ['jeberry308@gmail.com'])
    mail.attach_file("/home/lawscraper/reports/Booking Statement Report.xlsx")
    mail.send()
    print("sent mail!")


write_to_excel()