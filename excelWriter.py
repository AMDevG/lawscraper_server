import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import datetime

from django.core.mail import send_mail, EmailMessage
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pcos_site.settings')

def write_to_excel(detail_data):

    target_wb = Workbook()
    data_rows = ['Name', 'Arrest Date', 'Agency',
    			 'Address', 'City','State','Zipcode',
    			 'Date of Birth', 'Place of Birth', 'Arrest Age',
    			 'Cell Location', 'Booking Type', 'Offense',
    			 'Statute', 'Bond Assessed', 'Bond Amount Due',
    			 'Charge Status', 'Arrest Type',   'Offense 2',
                 'Statute 2', 'Bond Assessed 2', 'Bond Amount Due 2',
                 'Charge Status 2', 'Arrest Type 2', 'Offense 3',
                 'Statute 3', 'Bond Assessed 3', 'Bond Amount Due 3',
                 'Charge Status 3', 'Arrest Type 3', 'Offense 4',
                 'Statute 4', 'Bond Assessed 4', 'Bond Amount Due 4',
                 'Charge Status 4', 'Arrest Type 4',  'Offense 5',
                 'Statute 5',  'Bond Assessed 5', 'Bond Amount Due 5',
                 'Charge Status 5', 'Arrest Type 5',  "-"]

    target_ws = target_wb.create_sheet()
    target_ws.title = "Arrest Data"

    for i in range(1, 101):
        col_letter = get_column_letter(i)
        target_ws.column_dimensions[col_letter].width = 30

    for col in target_ws.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            cell.alignment = alignment_obj

    counter = 1
    for item in data_rows:
        target_ws.cell(row = 1, column = counter).value = item
        counter+=1

    col_counter = 1
    for key in detail_data.keys():
        col_counter +=1
        for i in range(1,len(data_rows)):
            if i == 1:
                try:
            	    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['name']
                except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 2:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['arrest_date']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 3:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['agency']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 4:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['address']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 5:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['city']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 6:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['state']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 7:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['zipcode']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 8:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['dob']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 9:
                try:
            	    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['pob']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 10:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['arrest_age']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 11:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['cell_location']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 12:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['booking_type']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 13:
                try:
                	target_ws.cell(row = col_counter , column = i ).value = detail_data[key]['offense']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 14:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['statute']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 15:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_assessed']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 16:
                try:
                	target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_amount_due']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 17:
                try:
            	    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['charge_status']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 18:
                try:
            	    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['arrest_type']
            	except:
            	    target_ws.cell(row = col_counter, column = i ).value = "-"


        ##### CHARGE 2 ######

            elif i == 19:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['offense2']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 20:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['statute2']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 21:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_assessed2']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 22:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_amount_due2']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 23:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['charge_status2']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 24:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['arrest_type2']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"

        ### CHARGE 3 #####

            elif i == 25:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['offense3']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 26:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['statute3']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 27:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_assessed3']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 28:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_amount_due3']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 29:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['charge_status3']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 30:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['arrest_type3']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"

            ##### CHARGE 4 ####
            elif i == 31:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['offense4']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 32:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['statute4']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 33:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_assessed4']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 34:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_amount_due4']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 35:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['charge_status4']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 36:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['arrest_type4']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"

        ##### CHARGE 5 #####

            elif i == 37:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['offense5']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 38:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['statute5']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 39:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_assessed5']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 40:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['bond_amount_due5']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 41:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['charge_status5']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"
            elif i == 42:
                try:
                    target_ws.cell(row = col_counter, column = i ).value = detail_data[key]['arrest_type5']
                except:
                    target_ws.cell(row = col_counter, column = i ).value = "-"


    day_day = datetime.date.today()
    date_day = day_day - datetime.timedelta(hours=6)
    date_day = str(date_day)
    workbook_name = "Booking Statement Report" + date_day + ".xlsx"

    del_sheet = target_wb.get_sheet_by_name('Sheet')
    target_wb.remove_sheet(del_sheet)

    target_wb.save("/home/lawscraper/reports/"+workbook_name)
    #target_wb.save("/Users/johnberry/Desktop/"+workbook_name)

    mail = EmailMessage("New Booking Report Statement", "", 'bprecosheet@gmail.com', ['jeberry308@gmail.com', 'js@dedicateddefense.com', 'ls@dedicateddefense.com'])
    #mail = EmailMessage("New Booking Report Statement", "", 'bprecosheet@gmail.com', ['jeberry308@gmail.com'])

    mail.attach_file("/home/lawscraper/reports/"+workbook_name)
    #mail.attach_file("/Users/johnberry/Desktop/"+workbook_name)

    mail.send()
    print("sent mail!")

