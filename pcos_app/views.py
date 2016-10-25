from django.shortcuts import render
from django.http import HttpResponse
from django.core.mail import send_mail, EmailMessage
import dryscrape
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
import openpyxl as xl
import sys


def home(request):
	write_to_excel()
	return HttpResponse("Done")

def getIDs():
	search_ids = []
	#Scrapes results returned from searching all bookings of day
	#and retrieves docket numbers.
	#Returns type list
	#search_ids holds all inmate numbers and is returned to caller
	sess = dryscrape.Session()
	sess.visit('http://www.pcsoweb.com/inmatebooking/')

	#Selects submit button, check include, and results per page
	#Include charge must be selected to avoid clicking error since
	#calendar dropdown covers button when selected

	dateInput = sess.at_xpath('//*[@id="txtBookingDate"]')
	dateInput.set('10/6/2016')

	check = sess.at_css('#chkIncludeCharge')
	button = sess.at_css('#btnSearch')

	return_size = sess.at_css('#drpPageSize')
	return_size.set('100')

	check.click()
	button.click()

	response = sess.body()
	#response is the javascript rendered html
	#after submital button is clicked

	soup = BeautifulSoup(response)

	#Searches for regular expression matching end of id attached
	#to inmate numbers
	inmate_numbers = soup.findAll("span", {"id" : re.compile('lblJMSNumber.*')})
	for num in inmate_numbers:
		search_ids.append(num.text)

	return search_ids


def get_id_detail():
	search_ids = getIDs()
	#Appends inmate number to url to retrieve detailed info
	#on inmate, this is the target page to scrape

	sess = dryscrape.Session()

	test_id = search_ids[0]	#Using first element as test id, will need to create loop

	base_url = 'http://www.pcsoweb.com/inmatebooking/SubjectResults.aspx?id='
	target_url = base_url + test_id

	sess.visit(target_url)
	response = sess.body()

	return response

def parseTarget():

	# detail_date is the main dictionary containing all data for report

	detail_date = {}
	results = []
	charge_rows = []
	charge_data = []

	target_html = get_id_detail()

	soup = BeautifulSoup(target_html)

	name = soup.find("span", {"id" : 'lblName1'}).text
	docket = soup.find("span", {"id" : 'lblDocket1'}).text
	arrest_date = soup.find("span", {"id" : 'lblArrestDate1'}).text
	agency = soup.find("span", {"id" : 'lblAgency'}).text
	address = soup.find("span", {"id" : 'lblAddress'}).text
	city = soup.find("span", {"id" : 'lblCity'}).text
	state = soup.find("span", {"id" : 'lblState'}).text
	zipcode = soup.find("span", {"id" : 'lblZipCode'}).text
	race = soup.find("span", {"id" : 'lblRace1'}).text
	sex = soup.find("span", {"id" : 'lblSex1'}).text
	dob = soup.find("span", {"id" : 'lblDOB1'}).text
	pob = soup.find("span", {"id" : 'lblPOB'}).text
	arrest_age = soup.find("span", {"id" : 'lblArrestAge'}).text
	eyes = soup.find("span", {"id" : 'lblEyes'}).text
	hair = soup.find("span", {"id" : 'lblHair'}).text
	complexion = soup.find("span", {"id" : 'lblComplexion'}).text
	height = soup.find("span", {"id" : 'lblHeight'}).text
	weight = soup.find("span", {"id" : 'lblWeight'}).text
	cell_location = soup.find("span", {"id" : 'CellLocation'}).text
	account_balance = soup.find("span", {"id" : 'lblAccountBalance'}).text
	spin = soup.find("span", {"id" : 'lblSPIN'}).text
	booking_type = soup.find("span", {"id" : 'lblBookingType'}).text
	alias = soup.find("span", {"id" : 'lblAKA'}).text


	#Charge table is separate from other booking data

	charge_table = soup.find("table", id='tblCharges')
	rows = charge_table.findAll('td')
	for row in rows:
		charge_data.append(row.text)

	#Iterates over every other because of headers
	for i in range(1, len(charge_data),2):
		if i == 1:
			detail_date['charge_number'] = charge_data[i]
		elif i == 3:
			detail_date['agency_report_number'] = charge_data[i]
		elif i == 5:
			detail_date['offense'] = charge_data[i]
		elif i == 7:
			detail_date['statute'] = charge_data[i]
		elif i == 9:
			detail_date['case_number'] = charge_data[i]
		elif i == 11:
			detail_date['bond_assessed'] = charge_data[i]
		elif i == 13:
			detail_date['bond_amount_due'] = charge_data[i]
		elif i == 15:
			detail_date['charge_status'] = charge_data[i]
		elif i == 17:
			detail_date['arrest_type'] = charge_data[i]
		elif i == 19:
			detail_date['obts'] = charge_data[i]

	detail_date['name'] = name
	detail_date['docket'] = docket
	detail_date['arrest_date'] = arrest_date
	detail_date['agency'] = agency
	detail_date['address'] = address
	detail_date['city'] = city
	detail_date['state'] = state
	detail_date['zipcode'] = zipcode
	detail_date['race'] = race
	detail_date['sex'] = sex
	detail_date['dob'] = dob
	detail_date['pob'] = pob
	detail_date['arrest_age'] = arrest_age
	detail_date['eyes'] = eyes
	detail_date['hair'] = hair
	detail_date['complexion'] = complexion
	detail_date['height'] = height
	detail_date['weight'] = weight
	detail_date['cell_location'] = cell_location
	detail_date['account_balance'] = account_balance
	detail_date['spin'] = spin
	detail_date['booking_type'] = booking_type
	detail_date['alias'] = alias

	return detail_date

def write_to_excel():

	data_rows = ['Name', 'Docket', 'Arrest Date', 'Agency',
				 'Address', 'City','State','Zipcode','Race',
				 'Sex', 'Date of Birth', 'Place of Birth', 'Arrest Age',
				 'Eye Color', 'Hair Color', 'Complexion', 'Height',
				 'Weight', 'Cell Location', 'Account Balance', 'SPIN', 'Booking Type',
				 'Alias', 'Charge Number', 'Agency Report Number', 'Offense',
				 'Statute', 'Case Number', 'Bond Assessed', 'Bond Amount Due',
				 'Charge Status', 'Arrest Type', 'OBTS']

	detail_data  = parseTarget()
	target_wb = Workbook()
	target_ws = target_wb.create_sheet(0)
	target_ws.title = detail_data['docket']

	target_ws.cell(row = 1, column = 1).value = data_rows[0]

	for i in range(1,len(data_rows)):
		target_ws.cell(row = i + 1, column = 1 ).value = data_rows[i]

		if i == 1:
			target_ws.cell(row = i, column = 2 ).value = detail_data['name']
		elif i == 2:
			target_ws.cell(row = i, column = 2 ).value = detail_data['docket']
		elif i == 3:
			target_ws.cell(row = i, column = 2 ).value = detail_data['arrest_date']
		elif i == 4:
			target_ws.cell(row = i, column = 2 ).value = detail_data['agency']
		elif i == 5:
			target_ws.cell(row = i, column = 2 ).value = detail_data['address']
		elif i == 6:
			target_ws.cell(row = i, column = 2 ).value = detail_data['city']
		elif i == 7:
			target_ws.cell(row = i, column = 2 ).value = detail_data['state']
		elif i == 8:
			target_ws.cell(row = i, column = 2 ).value = detail_data['zipcode']
		elif i == 9:
			target_ws.cell(row = i, column = 2 ).value = detail_data['race']
		elif i == 10:
			target_ws.cell(row = i, column = 2 ).value = detail_data['sex']
		elif i == 11:
			target_ws.cell(row = i, column = 2 ).value = detail_data['dob']
		elif i == 12:
			target_ws.cell(row = i, column = 2 ).value = detail_data['pob']
		elif i == 13:
			target_ws.cell(row = i , column = 2 ).value = detail_data['arrest_age']
		elif i == 14:
			target_ws.cell(row = i, column = 2 ).value = detail_data['eyes']
		elif i == 15:
			target_ws.cell(row = i, column = 2 ).value = detail_data['hair']
		elif i == 16:
			target_ws.cell(row = i, column = 2 ).value = detail_data['complexion']
		elif i == 17:
			target_ws.cell(row = i, column = 2 ).value = detail_data['height']
		elif i == 18:
			target_ws.cell(row = i, column = 2 ).value = detail_data['weight']
		elif i == 19:
			target_ws.cell(row = i, column = 2 ).value = detail_data['cell_location']
		elif i == 20:
			target_ws.cell(row = i, column = 2 ).value = detail_data['account_balance']
		elif i == 21:
			target_ws.cell(row = i, column = 2 ).value = detail_data['spin']
		elif i == 22:
			target_ws.cell(row = i, column = 2 ).value = detail_data['booking_type']
		elif i == 23:
			target_ws.cell(row = i, column = 2 ).value = detail_data['alias']
		elif i == 24:
			target_ws.cell(row = i, column = 2 ).value = detail_data['charge_number']
		elif i == 25:
			target_ws.cell(row = i, column = 2 ).value = detail_data['agency_report_number']
		elif i == 26:
			target_ws.cell(row = i, column = 2 ).value = detail_data['offense']
		elif i == 27:
			target_ws.cell(row = i, column = 2 ).value = detail_data['statute']
		elif i == 28:
			target_ws.cell(row = i, column = 2 ).value = detail_data['case_number']
		elif i == 29:
			target_ws.cell(row = i, column = 2 ).value = detail_data['bond_assessed']
		elif i == 30:
			target_ws.cell(row = i, column = 2 ).value = detail_data['bond_amount_due']
		elif i == 31:
			target_ws.cell(row = i, column = 2 ).value = detail_data['charge_status']
		elif i == 32:
			target_ws.cell(row = i, column = 2 ).value = detail_data['arrest_type']
		elif i == 33:
			target_ws.cell(row = i, column = 2 ).value = detail_data['obts']



			"""detail_date['charge_number'] = charge_data[i]
		elif i == 3:
			detail_date['agency_report_number'] = charge_data[i]
		elif i == 5:
			detail_date['offense'] = charge_data[i]
		elif i == 7:
			detail_date['statute'] = charge_data[i]
		elif i == 9:
			detail_date['case_number'] = charge_data[i]
		elif i == 11:
			detail_date['bond_assessed'] = charge_data[i]
		elif i == 13:
			detail_date['bond_amount_due'] = charge_data[i]
		elif i == 15:
			detail_date['charge_status'] = charge_data[i]
		elif i == 17:
			detail_date['arrest_type'] = charge_data[i]
		elif i == 19:
			detail_date['obts'] = charge_data[i]"""

	##HARDCODED FILE PATH WILL NEED TO POINT TO SERVER PATH
	workbook_name = "Booking Statement Report.xlsx"
	target_wb.save("/Users/johnberry/Desktop/"+workbook_name)
	mail = EmailMessage("Hello", "testemail", 'bprecosheet@gmail.com', ['jeberry308@gmail.com'])
	##HARDCODED FILE PATH WILL NEED TO POINT TO SERVER PATH
	mail.attach_file("/Users/johnberry/Desktop/Booking Statement Report.xlsx")
	mail.send()

